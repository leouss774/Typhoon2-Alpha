import openpyxl, hashlib, os
files = {
 "v2  (Referentiel_indicateurs_risk_engine_2.xlsx)":  "/mnt/user-data/uploads/Referentiel_indicateurs_risk_engine_2.xlsx",
 "v2_1(Referentiel_indicateurs_risk_engine_2_1.xlsx)":"/mnt/user-data/uploads/Referentiel_indicateurs_risk_engine_2_1.xlsx",
}
snap = {}
for label, p in files.items():
    wb = openpyxl.load_workbook(p, read_only=True)
    ws = wb['Référentiel indicateurs']
    ids = []
    for r in range(5, ws.max_row+1):
        v = ws.cell(r,1).value
        if v not in (None, ''):
            ids.append(str(v).strip())
    wm = wb['Méthodes & formules']
    meths = [str(wm.cell(r,1).value).strip() for r in range(5, wm.max_row+1)
             if wm.cell(r,1).value not in (None,'')]
    wj = wb['Journal des études']
    studies = sum(1 for r in range(5, wj.max_row+1) if wj.cell(r,2).value not in (None,''))
    snap[label] = dict(ids=ids, meths=meths, studies=studies,
                       sheets=wb.sheetnames,
                       md5=hashlib.md5(open(p,'rb').read()).hexdigest()[:12],
                       size=os.path.getsize(p))
    wb.close()

for k,v in snap.items():
    print("%s\n  md5=%s size=%d\n  indicateurs=%d  methodes=%d  etudes=%d  onglets=%d"
          % (k, v['md5'], v['size'], len(v['ids']), len(v['meths']), v['studies'], len(v['sheets'])))

a = snap["v2  (Referentiel_indicateurs_risk_engine_2.xlsx)"]
b = snap["v2_1(Referentiel_indicateurs_risk_engine_2_1.xlsx)"]
add_i = [i for i in b['ids'] if i not in set(a['ids'])]
rem_i = [i for i in a['ids'] if i not in set(b['ids'])]
add_m = [m for m in b['meths'] if m not in set(a['meths'])]
print("\nDELTA v2 -> v2_1")
print("  indicateurs AJOUTES (%d): %s" % (len(add_i), add_i))
print("  indicateurs RETIRES (%d): %s" % (len(rem_i), rem_i))
print("  methodes AJOUTEES  (%d): %s" % (len(add_m), add_m))
print("  onglets ajoutes: %s" % [s for s in b['sheets'] if s not in set(a['sheets'])])
print("\n  doublons d'ID dans v2  :", len(a['ids'])-len(set(a['ids'])))
print("  doublons d'ID dans v2_1:", len(b['ids'])-len(set(b['ids'])))
